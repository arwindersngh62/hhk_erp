from django.shortcuts import render,redirect
from supporters.models import supporter
from .forms import IndonForm,AnondonForm
from .models import verified_donation,misc
from django.utils import timezone
from datetime import datetime
from django.views.generic.list import ListView
from django.http import FileResponse,HttpResponseRedirect
from os import remove
import zipfile
from django.urls import reverse
from django.core.mail import EmailMessage
# Create your views here.
from reportlab.pdfgen import canvas
import os
from django.contrib.auth.decorators import login_required
from num2words import num2words
from . import forms
from openpyxl import Workbook
from usersreg.models import member
from django.core.paginator import Paginator

# function to create a pdf file
def hello(f_name,rec_no,name,amount,mode,date,donor_pan,don_address_a,don_address_b,don_address_c):
    print(f_name,rec_no,name,amount,mode,date,donor_pan)
    c = canvas.Canvas(f_name)
    c.drawInlineImage("recFormat.jpg", 40, 20,510,770)
    c.drawString(450,583,str(rec_no))
    c.drawString(210, 552, name)
    #c.drawString(475, 637, str(date_rec))
    c.drawString(150, 525, str(amount)   )
    c.drawString(230, 499, (str(num2words(amount))).capitalize()+"  only")
    if donor_pan:
        c.drawString(260, 453, str(donor_pan))
    if don_address_a:
        c.drawString(220, 426, str(don_address_a))
    if don_address_b:
        c.drawString(220, 416, str(don_address_b))
    if don_address_c:
        c.drawString(220, 406, str(don_address_c))
    c.drawString(270, 355, mode)
    c.drawString(260, 380, str(date))
    c.save()

# view that shows a three options of adding donations when add donation is clicked
# this can be removed by adding a 2nd tier menu to donations
@login_required
def add_don(request):
# this supplies the user with current user and fetched the member for that user
    mem = member.objects.get(user = request.user)
    return render(request,'donation/add_don.html',{"member":mem})

# view for individual donation
@login_required
def don_ind(request):
    mem = member.objects.get(user = request.user)
    if request.method=="GET":
        supp = supporter.objects.all()
        return render(request,'donation/don_ind.html',{"message":"0","supp":supp})
    else:
        obj = request.POST["supp_name"]
        phone = int(obj.split(",")[1])
        obj = supporter.objects.get(phone_no = phone)
        form = IndonForm()
        return render(request,'donation/don_ind.html',{"message":"hello","supporter":obj,'form':form,'member':mem,'mode':'add'})

curr_data = []
#view for bulk donations which just adds elements to the list of supporters
@login_required
def don_bulk(request):
    global curr_data
    mem = member.objects.get(user = request.user)
    supp = supporter.objects.all()
    month = timezone.localdate().month
    year = timezone.localdate().year
    if request.method == "GET":
#        try:
#            type = request.GET["typea"]
#            try:
#                if type == "si":
#                    supp= supporter.objects.filter(is_si = True)
#                elif type == "cons":
#                    supp = supporter.objects.filter(is_constant = True)
#                elif type == "res":
#                    supp = supporter.objects.filter(resources = True)
#                elif type == "spon":
#                    supp = supporter.objects.filter(is_sponsor = True)
#                elif type == "onetime":
#                    supp = supporter.objects.filter(is_one_time=True)
#                else:
#                    supp = supporter.objects.all()
#            except:
#                print("error inie")
#        except:

        request.session["notes"] = []
#        type='all'
#        paginator  = Paginator(supp,20)
#        page = request.GET.get("page")
#        supporters = paginator.get_page(page)
#    return render(request,'donation/don_bulk.html',{"supp":supporters,"month":month,"year":year,'member':mem,'type':type})
        return render(request,'donation/don_bulk.html',{'member':mem,'supp':supp,"month":month,"year":year})
    else:

        sup_to_send = []
        try:
            notes = request.session['notes']
            if request.POST["supp_name"] not in notes:
                notes.append(request.POST["supp_name"])
            request.session['notes'] = notes
            print(request.session["notes"])
        except:
            print("session notes are: none")
            request.session["notes"] = []
        print(request.session['notes'])
        for i in request.session['notes']:
            phone = int(i.split(",")[1])
            obj = supporter.objects.get(phone_no = phone)
            sup_to_send.append(obj)

        print("sup_to_Send")
        return render(request,'donation/don_bulk.html',{'member':mem,'supp':supp,'sup_to_send':sup_to_send,"month":month,"year":year})
# method for anonymous donation addition
@login_required
def don_anon(request):
    mem = member.objects.get(user = request.user)
    if request.method=="GET":
        form = AnondonForm()
        return render(request,'donation/don_anon.html',{"form":form,'member':mem})
    if request.method == "POST":
        form = AnondonForm(request.POST)
        amount = request.POST['amount']
        date = request.POST['date_donation']
        mode = request.POST['mode_donation']
        date_rec = timezone.now()
        anon_don = verified_donation(amount =  amount , date_donation = date , mode_donation = mode , anon = True,date_rec = date_rec)
        anon_name_set=False
        if 'anon_name' in form.changed_data:
            anon_name_set = True
            anon_don.anon_name = request.POST['anon_name']
        anon_don.save()
        anon_don.recipt_num = index_rec_no(anon_don.rec_no)
        anon_don.save()
        if anon_name_set:
            f_name = [str(anon_don.anon_name)+str(anon_don.recipt_num)+'.pdf']
        else:
            f_name = ['No Name -'+str(anon_don.recipt_num)+'.pdf']

        hello(f_name[0],anon_don.recipt_num,str(anon_don.anon_name),anon_don.amount,anon_don.mode_donation,anon_don.date_donation,"","","","")
        try:            #check if save is checked by user
            if request.POST['mail']:
                if len(request.POST['email_id'])>0:
                    send_mail_A([str(request.POST["email_id"])],f_name[0],str(anon_don.anon_name))
                else:
                    send_mail_A(['Kalamenrichinglives@gmail.com'],f_name[0],str(anon_don.anon_name))
        except:
            pass
        try:
            if request.POST["save"]=="True":
                view_val = 'added'
        except:
            os.remove(f_name[0])
            view_val = 'added_none'
            f_name = 'NONE'

        values = [[anon_don.recipt_num,anon_don.anon_name, anon_don.amount, anon_don.date_donation, anon_don.mode_donation,anon_don.date_rec,anon_don.migrated]]
        return render(request,'donation/don_added.html',{"values":values,"f_name":f_name,"view":view_val,'member':mem})

# method that actually adds donations to the database
@login_required
def don_added(request):
    print("inside don added")
    mem = member.objects.get(user = request.user)
    # if type is 0 then individual donation is to be added
    if request.POST["type"]=="0":
        supp = supporter.objects.get(pk = int(request.POST["pk"]))
        rec = verified_donation(member = supp ,
                                amount = request.POST["amount"],
                                date_donation = request.POST["date_donation"],
                                mode_donation = request.POST["mode_donation"],
                                date_rec = timezone.now())
        rec.save()
        rec.recipt_num  = index_rec_no(rec.rec_no)
        rec.save()
        f_name = [str(rec.member.name)+str(rec.recipt_num)+'.pdf']
        hello(f_name[0],rec.recipt_num,rec.member.name,rec.amount,rec.mode_donation,rec.date_donation,rec.member.pan_card,rec.member.address_a,rec.member.address_b, rec.member.address_c)
        try:
            if request.POST['mail']:
                if supp.email_id == None:
                    email = request.POST["email_id"]
                    if len(email)>0:
                        send_mail_A([email],f_name[0],rec.member.name)
                    else:
                        print('mail id not valid')
                else:
                    send_mail_A([supp.email_id],f_name[0],rec.member.name)
        except:
            pass

        try:            #check if save is checked by user
            if request.POST['save']:
                view_val = 'added'
                pass
        except:
            os.remove(f_name[0])
            view_val = 'added_none'
            f_name = 'NONE'
        values = [[rec.recipt_num,rec.member.name, rec.amount, rec.date_donation, rec.mode_donation,rec.date_rec,rec.migrated]]

    # if type is 1 then bulk donation is to be added
    if request.POST["type"]=="1":
        print("type one selected")
        abc = list(request.POST.keys())
        choices = []
        dates = []
        values = []
        f_name=[]
        saves = []
        mails = []
        for choice in abc:
            if 'choice' in choice:
                choices.append(choice.strip('choice'))
            if 'recipt' in choice:
                saves.append(choice.strip('recipt'))
            if 'mail' in choice:
                mails.append(choice.strip('mail'))
        for j in choices:
            dates.append(request.POST['date' + j])
        dates.sort(key=lambda date: datetime.strptime(date, '%Y-%m-%d'))
        sorted_choices = []
        for i in choices:
            date = dates.pop(0)
            for j in choices:
                if j not in sorted_choices:
                    if date == request.POST['date' + j]:
                        sorted_choices.append(j)
        for i in sorted_choices:
            member_id = (request.POST['choice'+i])
            mem_in = supporter.objects.get(pk=member_id)
            don = verified_donation(member=mem_in)
            amount = request.POST['amount' + i]
            don.amount = amount
            don.date_donation = request.POST['date'+i]
            don.date_rec = timezone.now()
            don.mode_donation = request.POST['mode'+i]
            don.save()
            don.recipt_num = index_rec_no(don.rec_no)
            don.save()
            f_name.append(str(don.member.name)+str(don.recipt_num)+'.pdf')
            hello(f_name[-1],don.recipt_num,don.member.name,don.amount,don.mode_donation,don.date_donation,don.member.pan_card,don.member.address_a,don.member.address_b,don.member.address_c)
            if i in mails:
                if mem_in.email_id == None:
                    send_mail_A(['info@harhathkalam.org'],f_name[-1],don.member.name)
                else:
                    send_mail_A([mem_in.email_id],f_name[-1],don.member.name)
            if i not in saves:
                os.remove(f_name[-1])
                f_name.pop()

            values.append([don.recipt_num,don.member.name, don.amount, don.date_donation, don.mode_donation, don.date_rec,don.migrated])
        if len(saves)>0:
            view_val = 'added'
        else:
            view_val = 'added_none'
    return render(request,'donation/don_added.html',{"values":values,"f_name":f_name,"view":view_val,'member':mem})

# view donations as list

class DonationListView(ListView):
    #user = self.request.user
    #mem = member.objects.get(user = user)
    model = verified_donation
    paginate_by = 50  # if pagination is desired

# download reciept
@login_required
def download_rec(request):
    allkeys =  list(request.POST.keys())
    files = []
    for i in allkeys:
        if 'file' in i:
            files.append(i)
    if len(files)>1:
        zipoj = zipfile.ZipFile('filled.zip', 'w')
        for file in files:
            zipoj.write(request.POST[file])
        zipoj.close()
        zip_file = open('filled.zip', 'rb')
    else:
        zip_file = open(request.POST['file1'], 'rb')
    #print(a)
    #f_name = request.POST['file']
    return FileResponse(zip_file)

# import donations
def import_don(request):
    mem = member.objects.get(user = request.user)
    if request.method == 'POST':
        form = forms.Uploadfiles(request.POST, request.FILES)
        if form.is_valid():
            file_data = request.FILES['file']
            fs = FileSystemStorage()
            if fs.exists('temp.xlsx'):
                fs.delete('temp.xlsx')
            fs.save('temp.xlsx', file_data)
        mapping={'rec_no':'0','member_phone':'1','anon_name':'2','amount':'3','date_donation':'4','mode_donation':'5','date_rec':'6','migrated':'8','anon':'10','mail_sent':'11'}
        import_worker(mapping)
        form=forms.Uploadfiles()
        return render(request, 'donation/import.html', {'form': form,'member':mem})
    else:
        form = forms.Uploadfiles()
        return render(request, 'donation/import.html', {'form':form,'member':mem})

#export donations
def export_donation(request):
    try:
        os.remove("exports/donation_export.xlsx")
    except:
        pass
    wb = Workbook()
    ws = wb.active
    ws.title = "Donations"
    ws_anon = wb.create_sheet("Amonymous")
    supp = verified_donation.objects.filter(anon = False)
    supp_anon = verified_donation.objects.filter(anon = True)
    ws.cell(column = 1,row = 1).value = "Sr.No"
    ws.cell(column = 2,row = 1).value = "Reciept No"
    ws.cell(column = 3,row = 1).value = "Member Name"
    ws.cell(column = 4,row = 1).value = "Member Phone Number"
    ws.cell(column = 5,row = 1).value = "Amount"
    ws.cell(column = 6,row = 1).value = "Date of Donation"
    ws.cell(column = 7,row = 1).value = "Mode of Donation"
    ws.cell(column = 8,row = 1).value = "Date of Recipt Generation"
    ws.cell(column = 9,row = 1).value = "Migrated"
    ws.cell(column = 10,row = 1).value = "Mail Sent"
    row_no = 2
    for entity in supp:
        ws.cell(column = 1,row = row_no).value = row_no
        ws.cell(column = 2,row = row_no).value = entity.recipt_num
        ws.cell(column = 3,row = row_no).value = entity.member.name
        ws.cell(column = 4,row = row_no).value = entity.member.phone_no
        ws.cell(column = 5,row = row_no).value = entity.amount
        ws.cell(column = 6,row = row_no).value = entity.date_donation
        ws.cell(column = 7,row = row_no).value = entity.mode_donation
        ws.cell(column = 8,row = row_no).value = entity.date_rec
        ws.cell(column = 9,row = row_no).value = entity.migrated
        ws.cell(column = 10,row = row_no).value = entity.mail_sent
        row_no+=1
        ws.cell(column = 1,row = 1).value = "Sr.No"
    ws_anon.cell(column = 2,row = 1).value = "Reciept No"
    ws_anon.cell(column = 3,row = 1).value = "Anonymous Name"
    ws_anon.cell(column = 4,row = 1).value = "Amount"
    ws_anon.cell(column = 5,row = 1).value = "Date of Donation"
    ws_anon.cell(column = 6,row = 1).value = "Mode of Donation"
    ws_anon.cell(column = 7,row = 1).value = "Date of Recipt Generation"
    ws_anon.cell(column = 8,row = 1).value = "Migrated"
    ws_anon.cell(column = 9,row = 1).value = "Mail Sent"
    row_no = 2
    for entity in supp_anon:
        ws_anon.cell(column = 1,row = row_no).value = row_no
        ws_anon.cell(column = 2,row = row_no).value = entity.recipt_num
        ws_anon.cell(column = 3,row = row_no).value = entity.anon_name
        ws_anon.cell(column = 4,row = row_no).value = entity.amount
        ws_anon.cell(column = 5,row = row_no).value = entity.date_donation
        ws_anon.cell(column = 6,row = row_no).value = entity.mode_donation
        ws_anon.cell(column = 7,row = row_no).value = entity.date_rec
        ws_anon.cell(column = 8,row = row_no).value = entity.migrated
        ws_anon.cell(column = 9,row = row_no).value = entity.mail_sent
        row_no+=1
    wb.save("exports/donation_export.xlsx")
    return FileResponse(open("exports/donation_export.xlsx","rb"))

#change prefix
def set_prefix(request):
    mem = member.objects.get(user = request.user)
    if request.method == 'GET':
        try:
            misc_obj = misc.objects.get(pk=1)
            value = misc_obj.rec_no_prefix
        except:
            value = 0
        form = forms.miscForm()
        return render(request,'donation/edit_misc.html',{'form':form,'value':value,'member':mem})
    if request.method == "POST":
        try:
            misc_obj = misc.objects.get(pk=1)
        except:
            misc_obj = misc(rec_no_prefix = 0)

        misc_obj.rec_no_prefix = request.POST['rec_no_prefix']
        misc_obj.save()
        print(request.POST['rec_no_prefix'])
        form = forms.miscForm()
        return render(request,'donation/edit_misc.html',{'form':form,'value':misc.objects.get(pk=1).rec_no_prefix,'member':mem})

# if you wish to resend mails
def resend_mails(request):
    mem = member.objects.get(user = request.user)
    if request.method == "GET":
        don = verified_donation.objects.filter(mail_sent = False)
        return render(request,'donation/resend_mail.html',{'don':don,'member':mem})
    if request.method == "POST":
        don = verified_donation.objects.filter(mail_sent = False)
        return render(request,'donation/resend_mail.html',{'don':don,'member':mem})

# edit any donation
def edit_don(request):
    mem = member.objects.get(user = request.user)
    if request.method == "GET":
        don = verified_donation.objects.all()
        return render(request,'donation/edit_don.html',{'don':don,'member':mem})
    if request.method == "POST":
        rec_no = request.POST["value"]
        don_data = verified_donation.objects.get(rec_no = rec_no)
        obj = don_data.member
        form = IndonForm({'amount':don_data.amount,'mode_donation':don_data.mode_donation})
        sup = supporter.objects.all()
        return render(request,'donation/don_ind.html',{"new_sup":sup,"message":"hello","supporter":obj,'form':form,'member':mem,'mode':'edit','rec':don_data})


# adding edited donation to db

def don_edited(request):
    mem =member.objects.get(user = request.user)
    if request.method == "POST":
        don = verified_donation.objects.get(rec_no = request.POST["rec_no"])
        try:
            don.amount = request.POST["amount"]
        except:
            pass
        try:
            don.mode_donation = request.POST["mode_donation"]
        except:
            pass
        try:
            don.date_donation = request.POST["date_donation"]
        except:
            pass
        don.save()
    return HttpResponseRedirect(reverse('donation:edit_don'))

# get recept num based on rec_no and prefix
def index_rec_no(value_no):
    misc_obj = misc.objects.get(pk=1)
    value = misc_obj.rec_no_prefix
    out = int(value_no)+int(value)
    return(out)

# method for importing
def import_worker(mapping):
    wb = load_workbook('media/temp.xlsx')
    ws=wb['Sheet1']
    rows = ws.rows
    for row in rows:
        try:
            phone = int(row[int(mapping['phone_no'])].value)
            name = (row[int(mapping['name'])].value)
            age = (row[int(mapping['age'])].value)
            gender = (row[int(mapping['gender'])].value)
            alt_phone_no = (row[int(mapping['alt_phone_no'])].value)
            email = (row[int(mapping['email'])].value)
            ref = (row[int(mapping['ref'])].value)
            city = (row[int(mapping['city'])].value)
            addr = (row[int(mapping['adress'])].value)
            sup = supporter(name=name, age=age, gender=gender,
                         alt_phone_no=alt_phone_no,
                         email_id=email, city=city,
                         Reference=ref,address_a=addr,
                         phone_no=phone)
            sup.save()
            print(name)
        except:
            continue

# method for sending mail
def send_mail_A(send_to,attach_file,name):
    print("Sending mail")
    msg = EmailMessage("Donation Receipt"," Dear "+str(name)+", \n\nYou just helped one dream get a life by sending one more child to school! \nThank You for being our torchbearer. \nPlease Find attached the receipt for your contribution.","Kalamenrichinglives@gmail.com",send_to)
    msg.attach_file(attach_file)
    msg.send()
    print("mail sent")